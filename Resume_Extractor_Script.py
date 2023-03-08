#Intial commit have only accuracy to email extraction phone no. extraction to a good extent while name extraction is poor.

import os #Library to excuete os related opeation say reading folder
import re #Library to compute regression in name, email and phone patterns
import glob # To read a particular files in hierarchical manner 
from PyPDF2 import PdfReader # To read pdfs
import docx2txt #To read docx/ Word files
from openpyxl import Workbook # to write xlsx files after extraction

# set the path of the directory containing the files
path = 'Resume/'

# to rename all the .doc files to .docx as process remains same thereafter
for filename in os.listdir(path):
    if filename.endswith('.doc'):
        os.rename(os.path.join(path, filename), os.path.join(path, filename[:-4] + '.docx'))

# Define regular expressions to search for name, phone number, and email address can be changed as per requirement
name_pattern = re.compile(r'^\s*([A-Za-z]+ [A-Za-z]+)\s*$')
phone_pattern = re.compile(r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b')
email_pattern = re.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b')



# Create a new workbook
wb = Workbook()

# Select the active worksheet
ws = wb.active

# Add headers
ws['A1'] = 'Name'
ws['B1'] = 'Phone'
ws['C1'] = 'Email'
ws['D1'] = 'Filename'

# Initialize row counter
row_num = 2

# Loop through all the files in the folder
for filename in os.listdir(path):
    if filename.endswith('.docx'):
        # Load the contents of the Word document into a string variable
        doc_path = os.path.join(path, filename)
        text = docx2txt.process(doc_path)

        # Search for name, phone number, and email address using regular expressions
        name_match = name_pattern.search(text)
        phone_match = phone_pattern.search(text)
        email_match = email_pattern.search(text)

        # Extract the name, phone number, and email address from the matches
        name = name_match.group(1) if name_match else None
        phone = phone_match.group() if phone_match else None
        email = email_match.group() if email_match else None

        # Write the data to the worksheet
        ws.cell(row=row_num, column=1, value=name)
        ws.cell(row=row_num, column=2, value=phone)
        ws.cell(row=row_num, column=3, value=email)
        ws.cell(row=row_num, column=4, value=filename)

    elif filename.endswith(".pdf"):
         
        # Open the PDF file
        with open(os.path.join(folder_path, filename), "rb") as pdf_file:
            # Read the PDF file
            pdf_data = PdfReader(pdf_file)
            text = pdf_data.pages[0].extract_text()
             
            
            # Extract name, phone number, and email
            name_match = name_regex.search(text)
            phone_match = phone_regex.search(text)
            email_match = email_regex.search(text)
             
             
            
            # Extract the name, phone number, and email address from the matches
            name = name_match.group(1) if name_match else None
            phone = phone_match.group() if phone_match else None
            email = email_match.group() if email_match else None

            # Write the data to the worksheet
            ws.cell(row=row_num, column=1, value=name)
            ws.cell(row=row_num, column=2, value=phone)
            ws.cell(row=row_num, column=3, value=email)
            ws.cell(row=row_num, column=4, value=filename)

    # Increment the row counter
    row_num += 1
    else:
        print("Some problem encountered please check for dependencies ,realtive paths or enviorment in place.")

# Save the workbook
wb.save('resume_data.xlsx')

# Excuetion successfull
print("A file with name resume_data.xlsx have saved in your current directory")
